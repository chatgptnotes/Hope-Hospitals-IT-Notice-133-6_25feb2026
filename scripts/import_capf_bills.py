#!/usr/bin/env python3
"""
Import CAPF IPD Final Bills from Excel into Supabase.
Usage: python3 scripts/import_capf_bills.py
"""

import os
import glob
import psycopg2
import openpyxl

DB = dict(
    host='db.zyzmasxhfofgarfgusxk.supabase.co',
    port=5432, dbname='postgres', user='postgres',
    password='2ECY2N28MDEAfaKy', sslmode='require'
)

BILL_FOLDER = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "24.02.2026-ALL BILL FOLDER", "CAPF", "2024", "IPD"
)

def clean(val):
    """Strip leading ': ' from Excel cell values."""
    if val is None:
        return ""
    s = str(val).strip()
    if s.startswith(":"):
        s = s[1:].strip()
    return s

def parse_date(val):
    """Convert 'DATE:-  DD/MM/YYYY' to 'DD-MM-YYYY'."""
    if not val:
        return ""
    return str(val).strip().replace("DATE:-", "").strip().replace("/", "-")

def find_total(ws):
    """Find TOTAL BILL AMOUNT row, return Col F value."""
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and "TOTAL BILL AMOUNT" in str(cell):
                amount = row[5]
                if amount is not None:
                    return float(amount)
    return 0.0

def extract_treatment_info(ws):
    """
    Extract treatment type (Surgical/Conservative) + medicine cost + investigation cost.
    Returns: (treatment_type_str, medicine_cost, investigation_cost)
    """
    treatment_type = ''
    medicine_cost = 0.0
    investigation_cost = 0.0

    for row in ws.iter_rows(values_only=True):
        col_a = str(row[0] or '').strip()
        col_b = str(row[1] or '').strip()
        col_f = row[5]

        # Treatment type is in column A, row 21 area
        if ('Surgical' in col_a or 'Conservative' in col_a) and not treatment_type:
            treatment_type = col_a

        # Medicine cost
        if 'Medicine Charges' in col_b and col_f:
            medicine_cost = float(col_f)

        # Investigation/Pathology cost
        if 'Pathology Charges' in col_b and col_f:
            investigation_cost = float(col_f)

    treatment_code = (
        f"{treatment_type} | Med: Rs.{medicine_cost:,.0f} | Inv: Rs.{investigation_cost:,.0f}"
    )
    return treatment_code

def extract_bill(filepath, department="IPD"):
    """Extract all transaction fields from one CAPF Final Bill xlsx."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    def row(n):
        return rows[n - 1] if n <= len(rows) else (None,) * 11

    date          = parse_date(row(3)[3])       # D3 — Bill date
    invoice_number = clean(row(4)[2])           # C4 — Bill No (invoice reference)
    patient_id    = clean(row(5)[2])            # C5 — Registration No
    patient_name  = clean(row(6)[2])            # C6 — Patient Name
    case_no       = clean(row(13)[2])           # C13 — Case No (admission_id)
    diagnosis     = clean(row(15)[2])           # C15 — Diagnosis

    total          = find_total(ws)
    treatment_code = extract_treatment_info(ws)
    narration      = f"{patient_name}: {diagnosis}"[:100]

    return {
        "transaction_date":  None,    # date patient paid — unknown until payment received
        "patient_id":        patient_id,
        "admission_id":      case_no,
        "invoice_number":    invoice_number,    # Bill No goes here
        "voucher_no":        "",                # Payment receipt — filled later when received
        "invoice_amount":    total,
        "discount_amount":   0,
        "net_bill_amount":   total,
        "cash_amount":       0,
        "other_mode_amount": 0,                 # Amount received from corporate — filled later
        "transaction_type":  "Receipt",
        "department":        department,
        "treatment_code":    treatment_code,
        "narration":         narration,
    }

def main():
    xlsx_files = sorted(glob.glob(os.path.join(BILL_FOLDER, "*.xlsx")))
    if not xlsx_files:
        print(f"No xlsx files found in:\n  {BILL_FOLDER}")
        return

    print(f"Found {len(xlsx_files)} bill(s) in CAPF/2024/IPD\n")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    total_amount = 0

    for filepath in xlsx_files:
        filename = os.path.basename(filepath)
        try:
            r = extract_bill(filepath, department="IPD")
            cur.execute('''
                INSERT INTO transactions
                  (transaction_date, patient_id, admission_id, invoice_number, voucher_no,
                   invoice_amount, discount_amount, net_bill_amount,
                   cash_amount, other_mode_amount, transaction_type,
                   department, treatment_code, narration)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ''', (
                r["transaction_date"], r["patient_id"], r["admission_id"],
                r["invoice_number"], r["voucher_no"],
                r["invoice_amount"], r["discount_amount"], r["net_bill_amount"],
                r["cash_amount"], r["other_mode_amount"], r["transaction_type"],
                r["department"], r["treatment_code"], r["narration"]
            ))
            total_amount += r["invoice_amount"]
            print(f"  OK  {filename:22s}  {r['patient_id']:15s}  Rs.{r['invoice_amount']:>10,.0f}  {r['invoice_number']}")
        except Exception as e:
            print(f"  ✗  {filename}: {e}")

    conn.commit()
    cur.close()
    conn.close()
    print(f"\nDone. Total inserted: Rs.{total_amount:,.0f}")
    print("Refresh http://localhost:3000")

if __name__ == "__main__":
    main()
