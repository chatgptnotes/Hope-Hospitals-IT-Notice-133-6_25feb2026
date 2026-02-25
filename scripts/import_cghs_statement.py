"""Import CGHS OLD PORTAL statement (Apr 2024 – Mar 2025) into Supabase."""

import openpyxl
import psycopg2
from datetime import datetime

EXCEL = '/Users/murali/Downloads/STATEMENT OF CGHS FROM 01 APRIL 2024 TO 31 MARCH 2025.xlsx'
SHEET = 'OLD PORTAL ORG'

DB = dict(
    host='db.zyzmasxhfofgarfgusxk.supabase.co',
    port=5432, dbname='postgres', user='postgres',
    password='2ECY2N28MDEAfaKy', sslmode='require'
)

INSERT_SQL = """
INSERT INTO transactions
  (transaction_date, patient_id, admission_id, voucher_no,
   invoice_amount, discount_amount, net_bill_amount,
   cash_amount, other_mode_amount, transaction_type,
   department, treatment_code, narration, invoice_number)
VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
"""


def parse_date(val):
    """Convert mixed date formats to DD-MM-YYYY string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%d-%m-%Y')
    s = str(val).strip()
    # Try DD/MM/YYYY
    for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).strftime('%d-%m-%Y')
        except ValueError:
            continue
    return s  # fallback: return as-is


def to_str(val):
    """Convert numeric Excel value to clean string (remove .0)."""
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def to_num(val):
    """Convert to number, treating N/A and None as 0."""
    if val is None or str(val).strip().upper() == 'N/A':
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb[SHEET]

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Get existing voucher_nos for dedup
    cur.execute("SELECT voucher_no FROM transactions WHERE voucher_no != ''")
    existing = {r[0] for r in cur.fetchall()}
    print(f"Existing voucher_nos in DB: {len(existing)}")

    inserted = 0
    skipped = 0
    errors = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        # Skip empty rows
        if all(c.value is None for c in row):
            continue

        row_num = row[0].row
        voucher = str(row[3].value or '').strip()

        # Dedup by voucher_no
        if voucher and voucher in existing:
            skipped += 1
            continue

        try:
            rec = (
                parse_date(row[0].value),          # transaction_date
                to_str(row[1].value),               # patient_id
                to_str(row[2].value),               # admission_id
                voucher,                             # voucher_no
                to_num(row[4].value),               # invoice_amount
                to_num(row[5].value),               # discount_amount
                to_num(row[6].value),               # net_bill_amount
                to_num(row[7].value),               # cash_amount
                to_num(row[8].value),               # other_mode_amount
                str(row[9].value or '').strip(),    # transaction_type
                str(row[10].value or '').strip(),   # department
                row[11].value,                       # treatment_code (NULL if empty)
                row[12].value,                       # narration (NULL if empty)
                None,                                # invoice_number (not in CGHS data)
            )
            cur.execute(INSERT_SQL, rec)
            existing.add(voucher)
            inserted += 1
        except Exception as e:
            errors.append(f"Row {row_num}: {e}")
            conn.rollback()
            continue

    conn.commit()

    # Verify
    cur.execute("SELECT COUNT(*) FROM transactions")
    total = cur.fetchone()[0]

    print(f"\nInserted: {inserted}")
    print(f"Skipped (duplicate): {skipped}")
    print(f"Errors: {len(errors)}")
    for e in errors[:5]:
        print(f"  {e}")
    print(f"\nTotal rows in DB: {total}")

    # Spot check
    cur.execute("SELECT transaction_date, patient_id, voucher_no, invoice_amount, department FROM transactions LIMIT 3")
    print("\nSample rows:")
    for r in cur.fetchall():
        print(f"  {r}")

    conn.close()


if __name__ == '__main__':
    main()
