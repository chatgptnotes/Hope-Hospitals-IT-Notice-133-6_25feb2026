#!/usr/bin/env python3
"""
Universal importer — reads ALL xlsx bills from ALL scheme folders
and inserts into Supabase transactions table.
Skips already-imported bills (checks by invoice_number).
"""

import os, glob, re
import psycopg2
import openpyxl

DB = dict(
    host='db.zyzmasxhfofgarfgusxk.supabase.co',
    port=5432, dbname='postgres', user='postgres',
    password='2ECY2N28MDEAfaKy', sslmode='require'
)

BASE = "/Users/murali/133(6)tan based case/24.02.2026-ALL BILL FOLDER"

# Already imported — skip CAPF/2024/IPD
SKIP_FOLDER = os.path.join(BASE, "CAPF", "2024", "IPD")

# ── helpers ────────────────────────────────────────────────────────────────

def clean(v):
    if v is None: return ''
    s = str(v).strip()
    # strip leading ': ' or ':'
    s = re.sub(r'^[:\s]+', '', s).strip()
    return s

def find_field(rows, *labels):
    """Search col B (index 1) for any label; return cleaned col C value."""
    for row in rows:
        col_b = str(row[1] or '').strip().upper()
        for lbl in labels:
            if lbl.upper() in col_b:
                val = row[2]
                if val is not None:
                    return clean(str(val))
    return ''

def find_date(rows):
    """Search cols C, D, E in first 6 rows for a date pattern."""
    for row in rows[:6]:
        for col_idx in [2, 3, 4]:  # C, D, E
            val = str(row[col_idx] if len(row) > col_idx else '')
            if not val or val == 'None': continue
            # Remove label prefix
            val = re.sub(r'DATE[:\-\s]+', '', val, flags=re.IGNORECASE).strip()
            # Match DD/MM/YYYY or DD-MM-YYYY
            m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', val)
            if m:
                return f"{m.group(1).zfill(2)}-{m.group(2).zfill(2)}-{m.group(3)}"
    # Also check col A row 3 (some MP Police 2025 bills)
    for row in rows[:4]:
        val = str(row[0] or '')
        m = re.search(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', val)
        if m:
            return f"{m.group(1).zfill(2)}-{m.group(2).zfill(2)}-{m.group(3)}"
    return ''

def find_total(ws):
    """Find TOTAL BILL AMOUNT row, return Col F (index 5) value."""
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell and 'TOTAL BILL AMOUNT' in str(cell).upper():
                amt = row[5]
                if amt is not None:
                    try: return float(amt)
                    except: pass
    return 0.0

def extract_treatment(ws):
    treatment_type = ''
    medicine = 0.0
    investigation = 0.0
    for row in ws.iter_rows(values_only=True):
        a = str(row[0] or '').strip()
        b = str(row[1] or '').strip()
        f = row[5] if len(row) > 5 else None
        if not treatment_type:
            if any(k in a for k in ['Surgical', 'Conservative', 'Medical Management']):
                treatment_type = a.strip()
            elif any(k in b for k in ['Conservative Treatment', 'Surgical Treatment', 'Conservative Medical']):
                treatment_type = b.strip()
        if 'Medicine Charges' in b and f:
            try: medicine = float(f)
            except: pass
        if 'Pathology Charges' in b and f:
            try: investigation = float(f)
            except: pass
    t = treatment_type or 'Treatment'
    return f"{t} | Med: Rs.{medicine:,.0f} | Inv: Rs.{investigation:,.0f}"

def get_scheme(path):
    p = path.upper()
    if 'MPPOLICE' in p: return 'MP Police'
    if 'MPKAY' in p: return 'MPKAY'
    if 'SECR' in p: return 'SECR'
    if 'CENTRAL RAILWAY' in p: return 'Central Railway'
    if 'CGHS' in p: return 'CGHS'
    if 'WCL' in p: return 'WCL'
    if 'CAPF' in p: return 'CAPF'
    if 'INSURANCE' in p or 'TPA' in p: return 'Insurance/TPA'
    return 'Other'

def get_department(path):
    p = path.upper()
    if '/OPD/' in p or '\\OPD\\' in p: return 'OPD'
    # CGHS 2024 root (no IPD/OPD subfolder) — check if it's OPD bill by content
    return 'IPD'

def extract_bill(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, values_only=True))

    bill_date    = find_date(rows)
    bill_no      = find_field(rows, 'BILL NO')
    reg_no       = find_field(rows, 'REGISTRATION NO')
    patient      = find_field(rows, 'NAME OF PATIENT')
    case_no      = find_field(rows, 'CASE NO', 'NEIS CODE', 'EMPLOYEE ID NO', 'EMPLOYEE CODE')
    diagnosis    = find_field(rows, 'DIAGNOSIS', 'DAIGNOSIS', 'CAUSE OF DEATH')
    total        = find_total(ws)
    treatment    = extract_treatment(ws)
    scheme       = get_scheme(filepath)
    dept         = get_department(filepath)

    # Detect OPD from content if folder doesn't say so
    for row in rows[:3]:
        if row[2] and 'OPD' in str(row[2]).upper():
            dept = 'OPD'
            break

    narration = f"{patient}: {diagnosis}"[:100]

    return {
        'transaction_date':  None,
        'invoice_number':    bill_no,
        'voucher_no':        '',
        'patient_id':        reg_no,
        'admission_id':      case_no or reg_no,
        'invoice_amount':    total,
        'discount_amount':   0,
        'net_bill_amount':   total,
        'cash_amount':       0,
        'other_mode_amount': 0,
        'transaction_type':  'Receipt',
        'department':        dept,
        'treatment_code':    treatment,
        'narration':         narration,
    }, scheme

# ── main ───────────────────────────────────────────────────────────────────

def main():
    # Collect all xlsx files except already-imported CAPF/2024/IPD
    all_files = sorted(glob.glob(os.path.join(BASE, '**', '*.xlsx'), recursive=True))
    files = [f for f in all_files if not f.startswith(SKIP_FOLDER)]

    print(f"Found {len(files)} xlsx files to import\n")

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Fetch already-imported invoice numbers to avoid duplicates
    cur.execute("SELECT invoice_number FROM transactions WHERE invoice_number IS NOT NULL")
    existing = {row[0] for row in cur.fetchall()}
    print(f"Already in DB: {len(existing)} bills\n")

    ok = skip = err = 0
    grand_total = 0.0

    INSERT = '''
        INSERT INTO transactions
          (transaction_date, patient_id, admission_id, invoice_number, voucher_no,
           invoice_amount, discount_amount, net_bill_amount,
           cash_amount, other_mode_amount, transaction_type,
           department, treatment_code, narration)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    '''

    for fp in files:
        fname = os.path.relpath(fp, BASE)
        try:
            r, scheme = extract_bill(fp)

            if not r['invoice_number']:
                print(f"  SKIP (no bill no)  {fname}")
                skip += 1
                continue

            if r['invoice_number'] in existing:
                print(f"  DUP               {fname}  [{r['invoice_number']}]")
                skip += 1
                continue

            cur.execute(INSERT, (
                r['transaction_date'], r['patient_id'], r['admission_id'],
                r['invoice_number'], r['voucher_no'],
                r['invoice_amount'], r['discount_amount'], r['net_bill_amount'],
                r['cash_amount'], r['other_mode_amount'], r['transaction_type'],
                r['department'], r['treatment_code'], r['narration']
            ))
            existing.add(r['invoice_number'])
            grand_total += r['invoice_amount']
            ok += 1
            print(f"  OK  [{scheme:15s}] {r['patient_id']:15s}  {r['invoice_number']:18s}  Rs.{r['invoice_amount']:>10,.0f}  {os.path.basename(fp)}")

        except Exception as e:
            err += 1
            print(f"  ERR {fname}: {e}")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\n{'='*70}")
    print(f"Done.  Inserted: {ok}  |  Skipped/Dup: {skip}  |  Errors: {err}")
    print(f"Grand total imported this run: Rs.{grand_total:,.0f}")
    print("Refresh http://localhost:3000")

if __name__ == '__main__':
    main()
